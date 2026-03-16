"""Экран отчётов: Общий / Покупатели / Поставщики — тёмная тема, фиолетовый акцент."""
from kivy.lang import Builder
from kivy.metrics import dp
from kivy.properties import StringProperty
from kivy.uix.widget import Widget
from kivymd.uix.screen import MDScreen
from kivymd.uix.card import MDCard
from kivymd.uix.button import MDRaisedButton, MDFlatButton
from kivymd.uix.boxlayout import MDBoxLayout
from kivymd.uix.label import MDLabel
from datetime import date as dt_date, timedelta
import db

Builder.load_string("""
#:import dp kivy.metrics.dp

<ReportsScreen>:
    md_bg_color: 0.12, 0.12, 0.14, 1

    MDFloatLayout:

        MDBoxLayout:
            orientation: "vertical"
            size_hint: 1, 1

            MDTopAppBar:
                title: "Отчеты"
                elevation: 0
                md_bg_color: 0.10, 0.10, 0.12, 1
                specific_text_color: 1, 1, 1, 1
                left_action_items: [["menu", lambda x: app.open_drawer()]]
                right_action_items: [["microsoft-excel", lambda x: root.export_excel()]]

            # Report type chips row
            MDBoxLayout:
                size_hint_y: None
                height: dp(38)
                md_bg_color: 0.12, 0.12, 0.14, 1
                padding: dp(8), dp(4)
                spacing: dp(6)
                id: type_row

            # Period chips row
            MDBoxLayout:
                size_hint_y: None
                height: dp(38)
                md_bg_color: 0.12, 0.12, 0.14, 1
                padding: dp(8), dp(4)
                spacing: dp(6)
                id: period_row

            ScrollView:
                MDBoxLayout:
                    orientation: "vertical"
                    adaptive_height: True
                    padding: dp(8)
                    spacing: dp(6)
                    id: report_content
""")

# ── helpers ────────────────────────────────────────────────

ACCENT = [0.60, 0.35, 0.85, 1]


def _fmt_money(val):
    s = f"{val:,.2f}".replace(",", " ").replace(".", ",")
    return f"{s}р."


def _fmt_date_header(date_str):
    months = {
        1: "янв.", 2: "фев.", 3: "мар.", 4: "апр.", 5: "мая", 6: "июн.",
        7: "июл.", 8: "авг.", 9: "сен.", 10: "окт.", 11: "нояб.", 12: "дек.",
    }
    try:
        d = dt_date.fromisoformat(date_str)
        return f"{d.day} {months[d.month]} {d.year} г."
    except Exception:
        return date_str


PERIOD_LABELS = {
    "today": "сегодня",
    "month": "месяц",
    "year": "год",
    "all": "всё время",
}


# ── screen ─────────────────────────────────────────────────

class ReportsScreen(MDScreen):
    report_type = StringProperty("general")
    period = StringProperty("today")

    # ── lifecycle ──────────────────────────────────────────
    def on_enter(self):
        self._build_type_chips()
        self._build_period_chips()
        self.load_report()

    def go_back(self):
        self.manager.current = "main_menu"

    # ── type / period setters ─────────────────────────────
    def set_type(self, t):
        self.report_type = t
        self._build_type_chips()
        self.load_report()

    def set_period(self, p):
        self.period = p
        self._build_period_chips()
        self.load_report()

    # ── chip builders ─────────────────────────────────────
    def _build_type_chips(self):
        row = self.ids.type_row
        row.clear_widgets()
        types = [("general", "Общий"), ("buyers", "Покупатели"), ("suppliers", "Поставщики")]
        for key, label in types:
            active = self.report_type == key
            if active:
                btn = MDRaisedButton(
                    text=label, font_size="12sp",
                    md_bg_color=ACCENT,
                    text_color=[1, 1, 1, 1],
                    size_hint_x=1,
                    on_release=lambda x, k=key: self.set_type(k),
                )
            else:
                btn = MDFlatButton(
                    text=label, font_size="12sp",
                    theme_text_color="Custom",
                    text_color=[0.6, 0.6, 0.6, 1],
                    size_hint_x=1,
                    on_release=lambda x, k=key: self.set_type(k),
                )
            row.add_widget(btn)

    def _build_period_chips(self):
        row = self.ids.period_row
        row.clear_widgets()
        periods = [("today", "Сегодня"), ("month", "Месяц"), ("year", "Год"), ("all", "Все")]
        for key, label in periods:
            active = self.period == key
            if active:
                btn = MDRaisedButton(
                    text=label, font_size="12sp",
                    md_bg_color=ACCENT,
                    text_color=[1, 1, 1, 1],
                    size_hint_x=1,
                    on_release=lambda x, k=key: self.set_period(k),
                )
            else:
                btn = MDFlatButton(
                    text=label, font_size="12sp",
                    theme_text_color="Custom",
                    text_color=[0.6, 0.6, 0.6, 1],
                    size_hint_x=1,
                    on_release=lambda x, k=key: self.set_period(k),
                )
            row.add_widget(btn)

    # ── date range ────────────────────────────────────────
    def _get_date_range(self):
        today = dt_date.today()
        if self.period == "today":
            s = today.strftime("%Y-%m-%d")
            return s, s
        elif self.period == "month":
            return today.replace(day=1).strftime("%Y-%m-%d"), today.strftime("%Y-%m-%d")
        elif self.period == "year":
            return today.replace(month=1, day=1).strftime("%Y-%m-%d"), today.strftime("%Y-%m-%d")
        return None, None

    # ── main dispatcher ───────────────────────────────────
    def load_report(self):
        if self.report_type == "general":
            self._general_report()
        elif self.report_type == "buyers":
            self._buyers_report()
        elif self.report_type == "suppliers":
            self._suppliers_report()

    # ── general report ────────────────────────────────────
    def _general_report(self):
        box = self.ids.report_content
        box.clear_widgets()

        d_from, d_to = self._get_date_range()
        if d_from:
            orders = db.get_orders_by_period(d_from, d_to)
            expenses = db.get_expenses_by_period(d_from, d_to)
        else:
            orders = db.get_orders_by_period("2000-01-01", "2099-12-31")
            expenses = db.get_expenses_by_period("2000-01-01", "2099-12-31")

        income_orders = [o for o in orders if o.get("doc_type") == "income"]
        expense_orders = [o for o in orders if o.get("doc_type") == "expense"]

        income_count = len(income_orders)
        expense_count = len(expense_orders)
        total_income = sum(o.get("total_sale", 0) for o in income_orders)
        total_expense = sum(o.get("total_sale", 0) for o in expense_orders)
        total_purchase = sum(o.get("total_purchase", 0) for o in expense_orders)
        total_profit = sum(o.get("profit", 0) for o in expense_orders)
        expenses_sum = sum(e.get("amount", 0) for e in expenses)
        net_profit = total_profit - expenses_sum

        # ── summary card ──
        period_label = PERIOD_LABELS.get(self.period, self.period)
        summary = (
            f"📊 Отчет за {period_label}\n\n"
            f"📦 Приходных: {income_count}  |  📦 Расходных: {expense_count}\n"
        )

        card = MDCard(
            orientation="vertical",
            size_hint_y=None,
            elevation=0,
            radius=[dp(12)],
            md_bg_color=[0.18, 0.18, 0.20, 1],
            padding=[dp(16), dp(12)],
        )
        card.bind(minimum_height=card.setter("height"))

        card.add_widget(MDLabel(
            text=summary,
            theme_text_color="Custom", text_color=[1, 1, 1, 1],
            font_style="Body1", size_hint_y=None, height=dp(70),
        ))
        card.add_widget(self._summary_line(
            "💰 Выручка:", total_income, [1, 1, 1, 1],
        ))
        card.add_widget(self._summary_line(
            "💳 Расход товара:", total_expense, [1, 1, 1, 1],
        ))
        card.add_widget(self._summary_line(
            "📈 Прибыль:", total_profit, [0.3, 0.85, 0.5, 1],
        ))
        card.add_widget(self._summary_line(
            "💸 Затраты:", expenses_sum, [0.95, 0.35, 0.35, 1],
        ))

        # separator
        sep = MDLabel(
            text="═══════════════",
            theme_text_color="Custom", text_color=[0.5, 0.5, 0.5, 1],
            size_hint_y=None, height=dp(24),
        )
        card.add_widget(sep)

        net_color = [0.3, 0.85, 0.5, 1] if net_profit >= 0 else [0.95, 0.35, 0.35, 1]
        card.add_widget(self._summary_line(
            "🏆 Чистая прибыль:", net_profit, net_color, bold=True,
        ))

        box.add_widget(card)

        # ── orders grouped by date ──
        by_date = {}
        for o in orders:
            by_date.setdefault(o["date"], []).append(o)

        for date_str in sorted(by_date.keys(), reverse=True):
            items = by_date[date_str]
            box.add_widget(MDLabel(
                text=_fmt_date_header(date_str),
                font_style="Subtitle2", bold=True,
                theme_text_color="Custom", text_color=[0.9, 0.9, 0.9, 1],
                size_hint_y=None, height=dp(32),
                padding=[dp(4), 0],
            ))
            for o in items:
                box.add_widget(self._make_order_card(o))

        if not orders and not expenses:
            box.add_widget(MDLabel(
                text="Нет данных за выбранный период",
                halign="center",
                theme_text_color="Custom", text_color=[0.5, 0.5, 0.5, 1],
                size_hint_y=None, height=dp(60),
            ))

        box.add_widget(Widget(size_hint_y=None, height=dp(80)))

    def _summary_line(self, label, value, color, bold=False):
        row = MDBoxLayout(
            orientation="horizontal",
            size_hint_y=None, height=dp(26),
        )
        row.add_widget(MDLabel(
            text=label, font_style="Body2",
            theme_text_color="Custom", text_color=[0.8, 0.8, 0.8, 1],
            size_hint_x=0.55,
        ))
        row.add_widget(MDLabel(
            text=_fmt_money(value),
            halign="right", font_style="Body1" if bold else "Body2",
            bold=bold,
            theme_text_color="Custom", text_color=color,
            size_hint_x=0.45,
        ))
        return row

    def _make_order_card(self, order):
        doc_type = order.get("doc_type", "")
        is_income = doc_type == "income"
        bar_color = [0.3, 0.85, 0.5, 1] if is_income else [0.95, 0.55, 0.15, 1]
        type_label = "Приход" if is_income else "Расход"

        card = MDCard(
            orientation="horizontal",
            size_hint_y=None, height=dp(72),
            elevation=0, radius=[dp(10)],
            md_bg_color=[0.18, 0.18, 0.20, 1],
            padding=[dp(12), dp(8)],
        )

        # colored left bar
        from kivy.graphics import Color as GColor, RoundedRectangle as GRR
        bar = Widget(size_hint=(None, 1), width=dp(4))
        with bar.canvas:
            GColor(*bar_color)
            GRR(pos=bar.pos, size=(dp(4), dp(50)), radius=[dp(2)])

        def _update_bar(instance, value):
            bar.canvas.clear()
            with bar.canvas:
                GColor(*bar_color)
                GRR(
                    pos=(instance.pos[0], instance.pos[1] + dp(10)),
                    size=(dp(4), instance.height - dp(20)),
                    radius=[dp(2)],
                )
        bar.bind(pos=_update_bar, size=_update_bar)
        card.add_widget(bar)

        # left info
        left = MDBoxLayout(orientation="vertical", padding=[dp(8), dp(4)], size_hint_x=0.6)
        doc_num = order.get("doc_number") or f"#{order.get('id', '')}"
        left.add_widget(MDLabel(
            text=f"{type_label} {doc_num}",
            font_style="Subtitle1", bold=True,
            theme_text_color="Custom", text_color=[1, 1, 1, 1],
            size_hint_y=0.5,
        ))
        left.add_widget(MDLabel(
            text=order.get("date", ""),
            font_style="Caption",
            theme_text_color="Custom", text_color=[0.6, 0.6, 0.6, 1],
            size_hint_y=0.5,
        ))
        card.add_widget(left)

        # right info
        right = MDBoxLayout(orientation="vertical", padding=[dp(4), dp(4)], size_hint_x=0.4)
        right.add_widget(MDLabel(
            text=_fmt_money(order.get("total_sale", 0)),
            halign="right", font_style="Subtitle2", bold=True,
            theme_text_color="Custom", text_color=[1, 1, 1, 1],
            size_hint_y=0.5,
        ))
        profit = order.get("profit", 0)
        p_color = [0.3, 0.85, 0.5, 1] if profit >= 0 else [0.95, 0.35, 0.35, 1]
        right.add_widget(MDLabel(
            text=_fmt_money(profit),
            halign="right", font_style="Caption",
            theme_text_color="Custom", text_color=p_color,
            size_hint_y=0.5,
        ))
        card.add_widget(right)
        return card

    # ── buyers report ─────────────────────────────────────
    def _buyers_report(self):
        box = self.ids.report_content
        box.clear_widgets()

        buyers = db.get_buyers()
        d_from, d_to = self._get_date_range()
        buyer_rows = []

        for b in buyers:
            all_orders = db.get_orders_by_buyer(b["id"])
            orders = self._filter_orders_by_period(all_orders, d_from, d_to)
            total_sale = sum(o.get("total_sale", 0) for o in orders)
            profit = sum(o.get("profit", 0) for o in orders)
            buyer_rows.append({
                "name": b["name"],
                "count": len(orders),
                "total_sale": total_sale,
                "profit": profit,
                "deposit": b.get("deposit", 0),
                "debt": b.get("debt", 0),
            })

        buyer_rows.sort(key=lambda r: r["total_sale"], reverse=True)

        if not buyer_rows:
            box.add_widget(MDLabel(
                text="Нет покупателей",
                halign="center",
                theme_text_color="Custom", text_color=[0.5, 0.5, 0.5, 1],
                size_hint_y=None, height=dp(60),
            ))
        else:
            for r in buyer_rows:
                box.add_widget(self._make_buyer_card(r))

        box.add_widget(Widget(size_hint_y=None, height=dp(80)))

    def _make_buyer_card(self, r):
        card = MDCard(
            orientation="vertical",
            size_hint_y=None,
            elevation=0, radius=[dp(12)],
            md_bg_color=[0.18, 0.18, 0.20, 1],
            padding=[dp(16), dp(12)],
        )
        card.bind(minimum_height=card.setter("height"))

        card.add_widget(MDLabel(
            text=f"👤 {r['name']}",
            font_style="Subtitle1", bold=True,
            theme_text_color="Custom", text_color=[1, 1, 1, 1],
            size_hint_y=None, height=dp(28),
        ))

        info_lines = [
            (f"Заказов: {r['count']}", [0.8, 0.8, 0.8, 1]),
            (f"Продажи: {_fmt_money(r['total_sale'])}", [1, 1, 1, 1]),
            (f"Прибыль: {_fmt_money(r['profit'])}", [0.3, 0.85, 0.5, 1]),
            (f"Депозит: {_fmt_money(r['deposit'])}", [0.6, 0.6, 0.6, 1]),
            (f"Долг: {_fmt_money(r['debt'])}", [0.95, 0.35, 0.35, 1] if r['debt'] > 0 else [0.6, 0.6, 0.6, 1]),
        ]
        for text, color in info_lines:
            card.add_widget(MDLabel(
                text=text, font_style="Body2",
                theme_text_color="Custom", text_color=color,
                size_hint_y=None, height=dp(22),
            ))
        return card

    # ── suppliers report ──────────────────────────────────
    def _suppliers_report(self):
        box = self.ids.report_content
        box.clear_widgets()

        suppliers = db.get_suppliers()
        d_from, d_to = self._get_date_range()
        supplier_rows = []

        for s in suppliers:
            all_orders = db.get_orders_by_supplier(s["id"])
            orders = self._filter_orders_by_period(all_orders, d_from, d_to)
            total_purchase = sum(o.get("total_purchase", 0) for o in orders)
            supplier_rows.append({
                "name": s["name"],
                "count": len(orders),
                "total_purchase": total_purchase,
            })

        supplier_rows.sort(key=lambda r: r["total_purchase"], reverse=True)

        if not supplier_rows:
            box.add_widget(MDLabel(
                text="Нет поставщиков",
                halign="center",
                theme_text_color="Custom", text_color=[0.5, 0.5, 0.5, 1],
                size_hint_y=None, height=dp(60),
            ))
        else:
            for r in supplier_rows:
                box.add_widget(self._make_supplier_card(r))

        box.add_widget(Widget(size_hint_y=None, height=dp(80)))

    def _make_supplier_card(self, r):
        card = MDCard(
            orientation="vertical",
            size_hint_y=None,
            elevation=0, radius=[dp(12)],
            md_bg_color=[0.18, 0.18, 0.20, 1],
            padding=[dp(16), dp(12)],
        )
        card.bind(minimum_height=card.setter("height"))

        card.add_widget(MDLabel(
            text=f"🏭 {r['name']}",
            font_style="Subtitle1", bold=True,
            theme_text_color="Custom", text_color=[1, 1, 1, 1],
            size_hint_y=None, height=dp(28),
        ))
        card.add_widget(MDLabel(
            text=f"Заказов: {r['count']}",
            font_style="Body2",
            theme_text_color="Custom", text_color=[0.8, 0.8, 0.8, 1],
            size_hint_y=None, height=dp(22),
        ))
        card.add_widget(MDLabel(
            text=f"Закупки: {_fmt_money(r['total_purchase'])}",
            font_style="Body2",
            theme_text_color="Custom", text_color=[1, 1, 1, 1],
            size_hint_y=None, height=dp(22),
        ))
        return card

    # ── helper: filter orders by date range ───────────────
    @staticmethod
    def _filter_orders_by_period(orders, d_from, d_to):
        if not d_from:
            return orders
        return [o for o in orders if d_from <= o.get("date", "") <= d_to]

    # ── Excel export ──────────────────────────────────────
    def export_excel(self):
        import os
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            from kivymd.uix.dialog import MDDialog
            MDDialog(title="openpyxl не установлен", text="pip install openpyxl").open()
            return

        d_from, d_to = self._get_date_range()
        if d_from:
            orders = db.get_orders_by_period(d_from, d_to)
            expenses = db.get_expenses_by_period(d_from, d_to)
        else:
            orders = db.get_orders_by_period("2000-01-01", "2099-12-31")
            expenses = db.get_expenses_by_period("2000-01-01", "2099-12-31")

        wb = Workbook()
        green_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
        header_font = Font(bold=True, size=11)
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        # Sheet 1: Summary
        ws = wb.active
        ws.title = "Отчет"
        period_label = PERIOD_LABELS.get(self.period, self.period)
        income_orders = [o for o in orders if o.get("doc_type") == "income"]
        expense_orders = [o for o in orders if o.get("doc_type") == "expense"]
        total_income = sum(o.get("total_sale", 0) for o in income_orders)
        total_expense = sum(o.get("total_sale", 0) for o in expense_orders)
        total_profit = sum(o.get("profit", 0) for o in expense_orders)
        expenses_sum = sum(e.get("amount", 0) for e in expenses)
        net_profit = total_profit - expenses_sum

        summary_data = [
            [f"Отчет за {period_label}"],
            [],
            ["Показатель", "Значение"],
            ["Приходных документов", len(income_orders)],
            ["Расходных документов", len(expense_orders)],
            ["Выручка", total_income],
            ["Расход товара", total_expense],
            ["Прибыль", total_profit],
            ["Затраты", expenses_sum],
            ["Чистая прибыль", net_profit],
        ]
        for row in summary_data:
            ws.append(row)
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 20

        # Sheet 2: Orders
        ws2 = wb.create_sheet("Документы")
        ws2.append(["Дата", "Номер", "Тип", "Закупка", "Продажа", "Прибыль"])
        for cell in ws2[1]:
            cell.font = header_font
            cell.fill = green_fill
            cell.border = thin_border
        for o in orders:
            ws2.append([
                o.get("date", ""),
                o.get("doc_number", ""),
                "Приход" if o.get("doc_type") == "income" else "Расход",
                o.get("total_purchase", 0),
                o.get("total_sale", 0),
                o.get("profit", 0),
            ])
        for col in ["A", "B", "C", "D", "E", "F"]:
            ws2.column_dimensions[col].width = 16

        # Sheet 3: Expenses
        if expenses:
            ws3 = wb.create_sheet("Затраты")
            ws3.append(["Дата", "Категория", "Описание", "Сумма"])
            for cell in ws3[1]:
                cell.font = header_font
                cell.fill = green_fill
                cell.border = thin_border
            for e in expenses:
                ws3.append([
                    e.get("date", ""),
                    e.get("category", ""),
                    e.get("description", ""),
                    e.get("amount", 0),
                ])
            for col in ["A", "B", "C", "D"]:
                ws3.column_dimensions[col].width = 20

        # Save
        today = dt_date.today().strftime("%Y-%m-%d")
        if hasattr(os, "environ") and "ANDROID_STORAGE" in os.environ:
            path = f"/storage/emulated/0/Download/report_{today}.xlsx"
        else:
            path = os.path.join(os.path.expanduser("~"), "Desktop", f"report_{today}.xlsx")
        wb.save(path)

        from kivymd.uix.dialog import MDDialog
        from kivymd.uix.button import MDFlatButton
        dlg = MDDialog(
            title="✅ Экспорт готов",
            text=f"Файл сохранён:\n{path}",
            buttons=[MDFlatButton(text="OK", on_release=lambda x: dlg.dismiss())],
        )
        dlg.open()
